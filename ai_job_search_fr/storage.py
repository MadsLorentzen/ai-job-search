"""Persistance locale, atomique et portable."""

from __future__ import annotations

import csv
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any, Iterable, Mapping
import zipfile
from xml.etree import ElementTree as ET


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_text(path: Path) -> tuple[str, str | None]:
    raw = path.read_bytes()
    warning = None
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        warning = "encodage non UTF-8 détecté; conversion avec remplacement des octets invalides"
        text = raw.decode("utf-8", errors="replace")
    if "�" in text and warning is None:
        warning = "caractères de remplacement présents dans le texte source"
    return text, warning


def _yaml_available():
    try:
        import yaml  # type: ignore
        return yaml
    except ImportError:
        return None


def load_yaml(path: Path) -> Any:
    text, _ = read_text(path)
    yaml = _yaml_available()
    if yaml:
        return yaml.safe_load(text) or {}
    # JSON est un sous-ensemble de YAML et constitue un fallback sans dépendance.
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return _load_simple_yaml(text)


def dump_yaml(data: Any) -> str:
    yaml = _yaml_available()
    if yaml:
        return yaml.safe_dump(data, allow_unicode=True, sort_keys=False, default_flow_style=False)
    # JSON indenté reste valide YAML 1.2 et évite d'exiger une dépendance.
    return json.dumps(data, ensure_ascii=False, indent=2) + "\n"


def write_yaml_atomic(path: Path, data: Any) -> None:
    write_text_atomic(path, dump_yaml(data))


def write_json_atomic(path: Path, data: Any) -> None:
    write_text_atomic(path, json.dumps(data, ensure_ascii=False, indent=2) + "\n")


def write_text_atomic(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="", dir=path.parent, delete=False) as handle:
        temp_name = Path(handle.name)
        handle.write(text)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_name, path)


def read_json(path: Path) -> Any:
    return json.loads(read_text(path)[0])


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    text, _ = read_text(path)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return [{str(k).strip(): (v or "").strip() for k, v in row.items() if k is not None} for row in reader]


def write_csv_atomic(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: list[str] | None = None) -> None:
    rows = list(rows)
    if fieldnames is None:
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(str(key))
        fieldnames = keys
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="", dir=path.parent, delete=False) as handle:
        temp_name = Path(handle.name)
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows({k: "" if row.get(k) is None else row.get(k) for k in fieldnames} for row in rows)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_name, path)


def extract_pdf_text(path: Path) -> tuple[str, str | None]:
    """Extrait le texte sans imposer de moteur PDF au projet."""
    pdftotext = shutil.which("pdftotext")
    if pdftotext:
        result = subprocess.run([pdftotext, "-layout", str(path), "-"], capture_output=True, text=True, encoding="utf-8", errors="replace", check=False)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout, None
    try:
        from pypdf import PdfReader  # type: ignore
        return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages), None
    except ImportError:
        return "", "PDF non extrait: installez pdftotext ou pypdf"
    except Exception as exc:  # pragma: no cover - dépend du PDF
        return "", f"échec extraction PDF: {exc}"


def read_xlsx_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, str]]:
    """Lit la première feuille XLSX avec la bibliothèque standard.

    openpyxl est utilisé s'il est disponible; sinon le lecteur minimal suffit
    pour les feuilles tabulaires du tracker (shared strings et cellules inline).
    """
    try:
        import openpyxl  # type: ignore
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(x or "").strip() for x in rows[0]]
        return [{headers[i]: "" if i >= len(row) or row[i] is None else str(row[i]).strip() for i in range(len(headers)) if headers[i]} for row in rows[1:]]
    except ImportError:
        pass
    with zipfile.ZipFile(path) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            for item in root.findall("x:si", ns):
                shared.append("".join(t.text or "" for t in item.findall(".//x:t", ns)))
        sheet_path = "xl/worksheets/sheet1.xml"
        if sheet_name:
            ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main", "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml")) if "xl/workbook.xml" in archive.namelist() else None
            relationships: dict[str, str] = {}
            if "xl/_rels/workbook.xml.rels" in archive.namelist():
                rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
                rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
                relationships = {rel.attrib.get("Id"): rel.attrib.get("Target", "") for rel in rel_root.findall("r:Relationship", rel_ns)}
            if workbook_root is not None:
                for sheet in workbook_root.findall("x:sheets/x:sheet", ns):
                    if sheet.attrib.get("name") == sheet_name:
                        target = relationships.get(sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"), "")
                        sheet_path = target.lstrip("/") if target else sheet_path
                        break
        if sheet_path not in archive.namelist():
            return []
        root = ET.fromstring(archive.read(sheet_path))
        ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        parsed: list[list[str]] = []
        for row in root.findall(".//x:sheetData/x:row", ns):
            cells: dict[int, str] = {}
            for cell in row.findall("x:c", ns):
                ref = cell.attrib.get("r", "A1")
                col = re.match(r"([A-Z]+)", ref).group(1)
                index = 0
                for char in col:
                    index = index * 26 + ord(char) - 64
                index -= 1
                value = cell.find("x:v", ns)
                text = value.text if value is not None and value.text is not None else ""
                if cell.attrib.get("t") == "s" and text.isdigit() and int(text) < len(shared):
                    text = shared[int(text)]
                inline = cell.find(".//x:t", ns)
                if inline is not None:
                    text = inline.text or ""
                cells[index] = text
            parsed.append([cells.get(i, "") for i in range(max(cells, default=-1) + 1)])
    if not parsed:
        return []
    headers = [str(x).strip() for x in parsed[0]]
    return [{headers[i]: row[i] if i < len(row) else "" for i in range(len(headers)) if headers[i]} for row in parsed[1:]]


def write_xlsx_atomic(path: Path, rows: Iterable[Mapping[str, Any]], fieldnames: list[str] | None = None) -> None:
    """Écrit une feuille XLSX tabulaire sans dépendance externe."""
    rows = list(rows)
    if fieldnames is None:
        fieldnames = []
        for row in rows:
            for key in row:
                if str(key) not in fieldnames:
                    fieldnames.append(str(key))
    def col(index: int) -> str:
        result = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
    def cell(ref: str, value: Any) -> str:
        text = "" if value is None else str(value)
        escaped = (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))
        return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escaped}</t></is></c>'
    row_xml = []
    all_rows = [{key: key for key in fieldnames}, *rows]
    for row_index, row in enumerate(all_rows, 1):
        row_xml.append(f'<row r="{row_index}">' + "".join(cell(f"{col(column_index)}{row_index}", row.get(key, "")) for column_index, key in enumerate(fieldnames, 1)) + "</row>")
    sheet = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>' + "".join(row_xml) + "</sheetData></worksheet>"
    files = {
        "[Content_Types].xml": '<?xml version="1.0" encoding="UTF-8"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>',
        "_rels/.rels": '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>',
        "xl/workbook.xml": '<?xml version="1.0" encoding="UTF-8"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Tracker" sheetId="1" r:id="rId1"/></sheets></workbook>',
        "xl/_rels/workbook.xml.rels": '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>',
        "xl/worksheets/sheet1.xml": sheet,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=path.parent, delete=False) as handle:
        temp_name = Path(handle.name)
    try:
        with zipfile.ZipFile(temp_name, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, content in files.items():
                archive.writestr(name, content.encode("utf-8"))
        os.replace(temp_name, path)
    finally:
        if temp_name.exists():
            temp_name.unlink()


def import_documents(paths: Iterable[Path], documents_dir: Path) -> dict[str, Any]:
    """Conserve les originaux et émet un journal idempotent."""
    documents_dir.mkdir(parents=True, exist_ok=True)
    journal_path = documents_dir / "import-journal.json"
    existing = read_json(journal_path) if journal_path.exists() else {"version": 1, "entries": []}
    by_digest = {e.get("sha256"): e for e in existing.get("entries", [])}
    by_source = {e.get("source"): e for e in existing.get("entries", [])}
    added, modified, ignored, warnings = [], [], [], []
    for source in paths:
        source = Path(source)
        if not source.exists() or not source.is_file():
            warnings.append(f"fichier introuvable: {source}")
            continue
        digest = sha256_file(source)
        if digest in by_digest:
            ignored.append({"path": str(source), "reason": "doublon", "sha256": digest})
            continue
        target = documents_dir / f"{digest[:12]}-{source.name}"
        shutil.copy2(source, target)
        warning = None
        if source.suffix.lower() in {".txt", ".md", ".markdown", ".tex", ".csv"}:
            _, warning = read_text(source)
        elif source.suffix.lower() == ".pdf":
            _, warning = extract_pdf_text(source)
        entry = {"source": str(source), "stored": str(target), "sha256": digest, "imported_at": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(), "warning": warning}
        if str(source) in by_source and by_source[str(source)].get("sha256") != digest:
            modified.append({"path": str(source), "previous_sha256": by_source[str(source)].get("sha256"), "sha256": digest})
        existing.setdefault("entries", []).append(entry)
        by_digest[digest] = entry
        added.append(entry)
        if warning:
            warnings.append(f"{source}: {warning}")
    write_json_atomic(journal_path, existing)
    return {"added": added, "modified": modified, "ignored": ignored, "warnings": warnings, "journal": str(journal_path)}


def _scalar(value: str) -> Any:
    value = value.strip()
    if not value:
        return None
    if value in {"null", "~"}:
        return None
    if value.lower() in {"true", "false"}:
        return value.lower() == "true"
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        return value[1:-1]
    if value.startswith("[") and value.endswith("]"):
        inner = value[1:-1].strip()
        if not inner:
            return []
        parts = re.split(r",\s*(?=(?:[^\"']|\"[^\"]*\"|'[^']*')*$)", inner)
        return [_scalar(part) for part in parts]
    if value.startswith("{") and value.endswith("}"):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value


def _load_simple_yaml(text: str) -> Any:
    """Petit parseur YAML pour les fichiers générés par ce projet."""
    lines = [line for line in text.splitlines() if line.strip() and not line.lstrip().startswith("#")]
    root: Any = {}
    stack: list[tuple[int, Any]] = [(-1, root)]
    for raw in lines:
        indent = len(raw) - len(raw.lstrip(" "))
        content = raw.strip()
        while stack and indent <= stack[-1][0]:
            stack.pop()
        parent = stack[-1][1]
        if content.startswith("- "):
            if isinstance(parent, list):
                item = content[2:].strip()
                if ":" in item and not item.startswith(("http://", "https://")):
                    key, value = item.split(":", 1)
                    obj = {key.strip(): _scalar(value)}
                    parent.append(obj)
                    stack.append((indent, obj))
                else:
                    parent.append(_scalar(item))
            continue
        if ":" not in content:
            continue
        key, value = content.split(":", 1)
        key = key.strip()
        value = value.strip()
        if value:
            if isinstance(parent, dict):
                parent[key] = _scalar(value)
        elif isinstance(parent, dict):
            # Examine the next meaningful line to choose list or mapping.
            idx = lines.index(raw)
            next_line = lines[idx + 1] if idx + 1 < len(lines) else ""
            child: Any = [] if next_line.strip().startswith("-") else {}
            parent[key] = child
            stack.append((indent, child))
    return root
