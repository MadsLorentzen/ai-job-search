#!/usr/bin/env python3
"""
Tests for Bug #10 fix in tools/convert_salary_excel.py.

Bug: ws[header_row] random-access fails silently in openpyxl read_only=True mode,
     producing no output. Fix: save header row cells during iter_rows scan.

Each test creates an Excel file, runs the converter, and asserts the output.
"""

import json
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).parent.parent
SCRIPT = REPO_ROOT / "tools" / "convert_salary_excel.py"

PASS = "\033[92mPASS\033[0m"
FAIL = "\033[91mFAIL\033[0m"

results = []


def run_converter(excel_path: Path, extra_args: list[str] = []) -> tuple[int, dict | None, str]:
    """Run the converter script, return (exit_code, parsed_json_or_None, stderr)."""
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as tf:
        out_path = Path(tf.name)

    cmd = [sys.executable, str(SCRIPT), str(excel_path), "--output", str(out_path)] + extra_args
    proc = subprocess.run(cmd, capture_output=True, text=True)

    parsed = None
    if out_path.exists() and out_path.stat().st_size > 0:
        try:
            parsed = json.loads(out_path.read_text())
        except json.JSONDecodeError:
            pass
        out_path.unlink()

    return proc.returncode, parsed, proc.stderr


def make_basic_excel(path: Path, header_row: int = 1):
    """Create a simple valid salary Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Optionally insert blank rows before header
    for _ in range(header_row - 1):
        ws.append([])

    ws.append(["Company", "City", "Antal", "Indeks"])
    ws.append(["Acme Corp", "Copenhagen", 120, 105.3])
    ws.append(["Beta ApS", "Aarhus", 45, 98.7])
    ws.append(["Gamma A/S", "Odense", 200, 112.0])
    wb.save(path)


def check(name: str, condition: bool, detail: str = ""):
    tag = PASS if condition else FAIL
    msg = f"  [{tag}] {name}"
    if detail:
        msg += f" — {detail}"
    print(msg)
    results.append((name, condition))


# ---------------------------------------------------------------------------
# Test 1: Basic happy path — header on row 1
# ---------------------------------------------------------------------------
def test_basic_happy_path():
    print("\nTest 1: Basic happy path (header on row 1)")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)
    make_basic_excel(p, header_row=1)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("JSON produced", data is not None)
    check("companies list non-empty", data is not None and len(data.get("companies", [])) > 0,
          f"got {len(data.get('companies', [])) if data else '?'} entries")
    check("correct company count", data is not None and len(data["companies"]) == 3,
          f"expected 3, got {len(data['companies']) if data else '?'}")
    check("first company name correct", data is not None and data["companies"][0]["company"] == "Acme Corp")
    check("city preserved", data is not None and data["companies"][1]["city"] == "Aarhus")


# ---------------------------------------------------------------------------
# Test 2: Header NOT on row 1 — should still find it (up to row 10)
# ---------------------------------------------------------------------------
def test_header_on_row_5():
    print("\nTest 2: Header on row 5 (blank rows above)")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)
    make_basic_excel(p, header_row=5)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("companies parsed", data is not None and len(data.get("companies", [])) == 3,
          f"expected 3, got {len(data['companies']) if data else '?'}")


# ---------------------------------------------------------------------------
# Test 3: Danish column headers (Firma / By)
# ---------------------------------------------------------------------------
def test_danish_headers():
    print("\nTest 3: Danish headers (Firma / By)")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Firma", "By", "Antal", "Indeks"])
    ws.append(["DanishCo", "Roskilde", 80, 103.5])
    wb.save(p)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("company parsed", data is not None and len(data["companies"]) == 1)
    check("company name correct", data is not None and data["companies"][0]["company"] == "DanishCo")


# ---------------------------------------------------------------------------
# Test 4: Missing company column — should warn and exit 1
# ---------------------------------------------------------------------------
def test_no_company_column():
    print("\nTest 4: No company column — expect exit 1 + warning")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Department", "Budget", "Headcount"])
    ws.append(["Engineering", 500000, 30])
    wb.save(p)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code non-zero", code != 0, f"got {code}")
    check("no JSON produced", data is None)
    check("warning in stderr", "Could not find" in stderr or "No data" in stderr, f"stderr: {stderr!r}")


# ---------------------------------------------------------------------------
# Test 5: Multiple sheets — all should be combined
# ---------------------------------------------------------------------------
def test_multiple_sheets():
    print("\nTest 5: Multiple sheets merged")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Region1"
    ws1.append(["Company", "City", "Indeks"])
    ws1.append(["AlphaX", "Copenhagen", 110.0])
    ws1.append(["BetaY", "Aarhus", 95.0])

    ws2 = wb.create_sheet("Region2")
    ws2.append(["Company", "City", "Indeks"])
    ws2.append(["GammaZ", "Odense", 102.5])

    wb.save(p)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("companies from both sheets", data is not None and len(data["companies"]) == 3,
          f"expected 3, got {len(data['companies']) if data else '?'}")


# ---------------------------------------------------------------------------
# Test 6: Empty rows interspersed — should skip them
# ---------------------------------------------------------------------------
def test_empty_data_rows_skipped():
    print("\nTest 6: Empty data rows skipped")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Company", "City", "Indeks"])
    ws.append(["ValidCo", "CPH", 100.0])
    ws.append([None, None, None])   # empty row — should skip
    ws.append(["AnotherCo", "ARH", 105.5])
    wb.save(p)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("only non-empty rows parsed", data is not None and len(data["companies"]) == 2,
          f"expected 2, got {len(data['companies']) if data else '?'}")


# ---------------------------------------------------------------------------
# Test 7: --source and --baseline flags pass through to metadata
# ---------------------------------------------------------------------------
def test_metadata_flags():
    print("\nTest 7: --source and --baseline metadata flags")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)
    make_basic_excel(p)

    code, data, stderr = run_converter(p, [
        "--source", "Union Stats 2025",
        "--baseline", "90",
        "--baseline-desc", "90 = lower quartile",
    ])
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("source in metadata", data is not None and data["metadata"]["source"] == "Union Stats 2025")
    check("baseline in metadata", data is not None and data["metadata"]["index_baseline"] == 90.0)
    check("baseline_description set", data is not None and "lower quartile" in data["metadata"]["baseline_description"])


# ---------------------------------------------------------------------------
# Test 8: Paired count+index columns grouped into categories correctly
# ---------------------------------------------------------------------------
# NOTE: detect_column_type uses "n" as a COUNT_PATTERNS substring, which is
# too broad — "n" appears in "indeks", "index", etc., misclassifying those
# as count. Use unambiguous headers ("Count"/"Salary") that bypass this
# pre-existing bug (separate from Bug #10).
# ---------------------------------------------------------------------------
def test_paired_count_index_columns():
    print("\nTest 8: Paired count/index columns grouped into categories")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)

    wb = openpyxl.Workbook()
    ws = wb.active
    # "Count" → COUNT_PATTERNS["count"] match; "Salary" → INDEX_PATTERNS["salary"] match
    # Both are unambiguous — no "n" substring collision
    ws.append(["Company", "City", "Count", "Salary"])
    ws.append(["PairCo", "Vejle", 200, 108.0])
    wb.save(p)

    code, data, stderr = run_converter(p)
    p.unlink()

    check("exit code 0", code == 0, f"got {code}\n{stderr}")
    check("company parsed", data is not None and len(data["companies"]) == 1)
    if data and data["companies"]:
        cats = data["companies"][0]["categories"]
        check("categories dict non-empty", len(cats) > 0, f"got: {list(cats.keys())}")
        # At least one category should have both count and index
        has_pair = any(
            isinstance(v, dict) and "count" in v and "index" in v
            for v in cats.values()
        )
        check("at least one paired category", has_pair, f"categories: {cats}")


# ---------------------------------------------------------------------------
# Test 9: File not found — should exit 1 with error message
# ---------------------------------------------------------------------------
def test_file_not_found():
    print("\nTest 9: File not found — expect exit 1")
    p = Path("/tmp/nonexistent_file_xyz_12345.xlsx")
    code, data, stderr = run_converter(p)
    check("exit code 1", code == 1, f"got {code}")
    check("error in stderr", "not found" in stderr.lower() or "error" in stderr.lower(), f"stderr: {stderr!r}")


# ---------------------------------------------------------------------------
# Test 10: THE CORE BUG — verify fix works: read_only=True must not produce empty headers
# ---------------------------------------------------------------------------
def test_read_only_mode_headers_not_empty():
    """
    Directly tests that parse_sheet populates headers when the workbook
    is opened read_only=True. This is the exact scenario that was broken.
    """
    print("\nTest 10: Core bug — parse_sheet works with read_only=True workbook")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        p = Path(tf.name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Company", "City", "Indeks"])
    ws.append(["ReadOnlyCo", "Esbjerg", 99.9])
    wb.save(p)

    # Open in read_only mode like the production code does, import parse_sheet directly
    sys.path.insert(0, str(REPO_ROOT / "tools"))
    from convert_salary_excel import parse_sheet  # type: ignore
    wb_ro = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws_ro = wb_ro[wb_ro.sheetnames[0]]
    companies = parse_sheet(ws_ro)
    wb_ro.close()
    p.unlink()

    check("parse_sheet returns non-empty list", len(companies) == 1,
          f"expected 1, got {len(companies)}")
    if companies:
        check("company name correct", companies[0]["company"] == "ReadOnlyCo",
              f"got: {companies[0]['company']!r}")
        check("categories populated (not empty dict)", len(companies[0]["categories"]) > 0,
              f"got: {companies[0]['categories']}")


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    test_basic_happy_path()
    test_header_on_row_5()
    test_danish_headers()
    test_no_company_column()
    test_multiple_sheets()
    test_empty_data_rows_skipped()
    test_metadata_flags()
    test_paired_count_index_columns()
    test_file_not_found()
    test_read_only_mode_headers_not_empty()

    total = len(results)
    passed = sum(1 for _, ok in results if ok)
    failed = total - passed

    print(f"\n{'='*50}")
    print(f"Results: {passed}/{total} passed", end="")
    if failed:
        print(f"  ({failed} FAILED)")
        for name, ok in results:
            if not ok:
                print(f"  - {name}")
    else:
        print(" — all green")

    sys.exit(0 if failed == 0 else 1)
